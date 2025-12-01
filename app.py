from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
import os
from datetime import datetime, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import desc
from typing import Optional, List
from io import BytesIO, StringIO
from flask import Flask, render_template
import smtplib
from email.message import EmailMessage

app = Flask(__name__)

# ============ 加载环境变量 ============
try:
    from dotenv import load_dotenv
    load_dotenv()  # 加载 .env 文件中的环境变量
    print("✅ 环境变量加载成功")
except ImportError:
    print("⚠️  python-dotenv 未安装，跳过环境变量加载")

# ============ 安全初始化应用 ============
app = Flask(__name__)

# 🔐 安全密钥配置 - 支持开发和生成环境
app.secret_key = os.environ.get("SECRET_KEY")
if not app.secret_key:
    if os.environ.get("FLASK_ENV") == "production":
        raise ValueError("❌ SECRET_KEY environment variable is required for production")
    else:
        # 开发环境使用一个默认密钥（不要在生产环境使用！）
        app.secret_key = "dev-secret-key-for-local-development-only-123456"
        print("⚠️  使用开发环境密钥，生产环境请设置 SECRET_KEY 环境变量")

# ============ 安全数据库配置 ============
# 从环境变量获取数据库URL
database_url = os.environ.get("DATABASE_URL")
if not database_url:
    # 开发环境回退到硬编码的数据库URL
    database_url = "postgresql://paid_user:zgevvYEGo2MaqkEjoC3LOdid5esaFSM7@dpg-d4dj33ndiees73ckpk3g-a.singapore-postgres.render.com/paid_website"
    print("⚠️  使用默认数据库URL，生产环境请设置 DATABASE_URL 环境变量")

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_recycle': 300,           # 5分钟回收连接
    'pool_pre_ping': True,         # 连接前检查
    'pool_size': 10,               # 连接池大小
    'max_overflow': 20,            # 最大溢出连接数
    'pool_timeout': 30,            # 获取连接超时时间
}

# 🔐 会话安全配置
app.config.update(
    SESSION_COOKIE_SECURE=True,      # 仅HTTPS传输
    SESSION_COOKIE_HTTPONLY=True,    # 防止XSS读取
    SESSION_COOKIE_SAMESITE='Lax',   # CSRF保护
    PERMANENT_SESSION_LIFETIME=timedelta(hours=1)  # 会话1小时后过期
)

db = SQLAlchemy(app)

# ============ 配置常量 ============
MEMBERSHIP_PRICE = 00
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "942521233@qq.com")  # 从环境变量获取

# 默认的内容分类和模块（防止未定义错误），可以根据实际内容替换为数据库或配置加载
CONTENT_CATEGORIES = [
    {"id": 1, "name": "入门指南", "slug": "getting-started"},
    {"id": 2, "name": "高级技巧", "slug": "advanced"},
    {"id": 3, "name": "常见问题", "slug": "faq"}
]

CONTENT_MODULES = [
    {"id": 1, "category_id": 1, "title": "如何注册与登录", "content": "在此处添加内容摘要..."},
    {"id": 2, "category_id": 1, "title": "支付流程说明", "content": "在此处添加支付流程..."},
    {"id": 3, "category_id": 2, "title": "优化技巧", "content": "在此处添加高级技巧..."},
]

def send_reset_email(user_email, reset_url):
    """使用替代方法的邮件发送函数"""
    try:
        # 获取邮件配置
        smtp_server = os.environ.get('MAIL_SERVER', '')
        smtp_port_str = os.environ.get('MAIL_PORT', '587')
        smtp_user = os.environ.get('MAIL_USERNAME', '')
        smtp_pass = os.environ.get('MAIL_PASSWORD', '')
        
        # 检查必要的配置是否存在
        if not smtp_server or not smtp_user or not smtp_pass:
            print("❌ 邮件配置不完整，无法发送邮件")
            return False
        
        # 确保端口是整数
        try:
            smtp_port = int(smtp_port_str)
        except (ValueError, TypeError):
            smtp_port = 587
        
        # 创建邮件消息
        msg = EmailMessage()
        msg['Subject'] = '上岸翻身营 - 密码重置'
        msg['From'] = f'上岸翻身营 <{smtp_user}>'
        msg['To'] = user_email
        
        # 邮件内容
        body = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <h3 style="color: #4361ee;">上岸翻身营 - 密码重置</h3>
            <p>您请求重置密码，请点击以下链接：</p>
            <p>
                <a href="{reset_url}" style="background: #4361ee; color: white; padding: 12px 24px; text-decoration: none; border-radius: 5px; display: inline-block;">
                    重置密码
                </a>
            </p>
            <p>或者复制以下链接到浏览器：</p>
            <p style="background: #f8f9fa; padding: 10px; border-radius: 5px; word-break: break-all;">
                {reset_url}
            </p>
            <p><strong>该链接1小时内有效。</strong></p>
            <p style="color: #6c757d; font-size: 14px;">
                如果不是您本人操作，请忽略此邮件。
            </p>
        </div>
        """
        
        msg.set_content(body, subtype='html')
        
        print(f"📧 尝试发送邮件到: {user_email}")
        
        # 发送邮件
        if smtp_port == 465:
            server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        else:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
        
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)
        server.quit()
        
        print(f"✅ 密码重置邮件已发送至: {user_email}")
        return True
        
    except Exception as e:
        print(f"❌ 邮件发送失败: {str(e)}")
        return False

# ============ 数据模型（兼容版本） ============
class User(db.Model):
    __tablename__ = "user"
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password = db.Column(db.String(200), nullable=False)
    is_admin = db.Column(db.Boolean, default=False, nullable=False)
    create_time = db.Column(db.DateTime, default=datetime.now, nullable=False, index=True)
    
    # 关系
    payments = db.relationship('Payment', backref='user', lazy=True, cascade='all, delete-orphan')
    questions = db.relationship('Question', backref='user', lazy=True, cascade='all, delete-orphan')
    
    def __init__(self, username: str, email: str, password: str, is_admin: bool = False):
        self.username = username
        self.email = email
        self.password = password
        self.is_admin = is_admin

class Payment(db.Model):
    __tablename__ = "payment"
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    amount = db.Column(db.Float, nullable=False)
    payment_method = db.Column(db.String(50), nullable=False)
    status = db.Column(db.String(20), default='pending', nullable=False, index=True)
    create_time = db.Column(db.DateTime, default=datetime.now, nullable=False, index=True)
    process_time = db.Column(db.DateTime)
    notes = db.Column(db.Text)
    
    def __init__(self, user_id: int, amount: float, payment_method: str, 
                 status: str = 'pending', notes: Optional[str] = None):
        self.user_id = user_id
        self.amount = amount
        self.payment_method = payment_method
        self.status = status
        self.notes = notes

class Question(db.Model):
    __tablename__ = "question"
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    content = db.Column(db.Text, nullable=False)
    answer = db.Column(db.Text)
    answered = db.Column(db.Boolean, default=False, nullable=False, index=True)
    create_time = db.Column(db.DateTime, default=datetime.now, nullable=False, index=True)
    answer_time = db.Column(db.DateTime)
    
    def __init__(self, user_id: int, content: str, answer: Optional[str] = None, 
                 answered: bool = False):
        self.user_id = user_id
        self.content = content
        self.answer = answer
        self.answered = answered

class PasswordReset(db.Model):
    """密码重置令牌模型"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    token = db.Column(db.String(100), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    expires_at = db.Column(db.DateTime, nullable=False)
    used = db.Column(db.Boolean, default=False)
    
    def __init__(self, user_id: int, token: str, expires_at: datetime):
        self.user_id = user_id
        self.token = token
        self.expires_at = expires_at

    def is_valid(self):
        """检查令牌是否有效"""
        return not self.used and self.expires_at > datetime.now()
    
class AdminLog(db.Model):
    __tablename__ = "admin_log"
    
    id = db.Column(db.Integer, primary_key=True)
    admin_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action = db.Column(db.String(200), nullable=False)
    target_type = db.Column(db.String(50), nullable=True)  # 改为 nullable=True
    target_id = db.Column(db.Integer, nullable=True)       # 改为 nullable=True
    ip_address = db.Column(db.String(45), nullable=True)   # 改为 nullable=True
    user_agent = db.Column(db.Text, nullable=True)         # 改为 nullable=True
    create_time = db.Column(db.DateTime, default=datetime.now, nullable=False)
    
    # 添加构造函数
    def __init__(self, admin_id, action, target_type=None, target_id=None, 
                 ip_address=None, user_agent=None):
        self.admin_id = admin_id
        self.action = action
        self.target_type = target_type
        self.target_id = target_id
        self.ip_address = ip_address
        self.user_agent = user_agent

# ============ 优化查询方法 ============

def get_pending_payments() -> List[Payment]:  # 现在 List 已导入
    """优化：获取待审核支付（使用索引）"""
    return db.session.query(Payment).filter_by(status='pending')\
                       .order_by(Payment.create_time.asc())\
                       .options(db.joinedload(Payment.user))\
                       .all()

def get_unanswered_questions() -> List[Question]:
    """优化：获取未回答问题（使用索引）"""
    return db.session.query(Question).filter_by(answered=False)\
                        .order_by(Question.create_time.asc())\
                        .options(db.joinedload(Question.user))\
                        .all()

def get_user_questions(user_id: int) -> List[Question]:
    """优化：获取用户问题列表"""
    return db.session.query(Question).filter_by(user_id=user_id)\
                        .order_by(Question.create_time.desc())\
                        .all()

# ============ 装饰器 ============
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('请先登录', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            flash('请先登录管理员账号', 'warning')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def payment_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        
        # 检查用户是否有已通过的支付
        approved_payment = Payment.query.filter_by(
            user_id=session['user_id'], 
            status='approved'
        ).first()
        
        if not approved_payment:
            flash('请先完成支付验证', 'warning')
            return redirect(url_for('payment_manual'))
        return f(*args, **kwargs)
    return decorated_function

# ============ 模板测试 ============
@app.template_test('date_equal')
def date_equal_test(dt, date_str_or_date):
    """日期相等判断测试"""
    if not isinstance(dt, datetime):
        return False
    
    # 处理 date_str_or_date 参数，可能是字符串或 date 对象
    if isinstance(date_str_or_date, str):
        # 如果是字符串，解析为日期
        compare_date = datetime.strptime(date_str_or_date, '%Y-%m-%d').date()
    else:
        # 如果已经是 date 对象，直接使用
        compare_date = date_str_or_date
    
    return dt.date() == compare_date

@app.template_test('date_ge')
def date_ge_test(dt, date_str_or_date):
    """日期大于等于判断测试"""
    if not isinstance(dt, datetime):
        return False
    
    # 处理 date_str_or_date 参数，可能是字符串或 date 对象
    if isinstance(date_str_or_date, str):
        # 如果是字符串，解析为日期
        compare_date = datetime.strptime(date_str_or_date, '%Y-%m-%d').date()
    else:
        # 如果已经是 date 对象，直接使用
        compare_date = date_str_or_date
    
    return dt.date() >= compare_date

# ============ 用户路由 ============
@app.route('/')
def index():
    """首页"""
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    """用户注册"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()

        # 新增：检查条款同意
        agree_terms = request.form.get('agree_terms')
        age_confirm = request.form.get('age_confirm')
        
        if not all([username, email, password]):
            return render_template('register.html', error="请填写所有字段")
        
        # 新增：检查是否同意条款
        if not agree_terms or not age_confirm:
            return render_template('register.html', error="请阅读并同意服务条款，并确认年龄要求")
        
        # 检查用户是否已存在
        if User.query.filter_by(username=username).first():
            return render_template('register.html', error="用户名已存在")
        if User.query.filter_by(email=email).first():
            return render_template('register.html', error="邮箱已被注册")
        
        try:
            new_user = User(
                username=username,
                email=email,
                password=generate_password_hash(password)
            )
            db.session.add(new_user)
            db.session.commit()
            
            # 设置会话
            session['user_id'] = new_user.id
            session['username'] = new_user.username
            session['email'] = new_user.email
            
            flash('注册成功！请完成支付验证', 'success')
            return redirect(url_for('payment_manual'))
            
        except Exception as e:
            db.session.rollback()
            return render_template('register.html', error=f"注册失败: {str(e)}")
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """用户登录"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if not all([username, password]):
            return render_template('login.html', error="请填写用户名和密码")
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['email'] = user.email
            
            # 检查支付状态
            approved_payment = Payment.query.filter_by(
                user_id=user.id, 
                status='approved'
            ).first()
            
            if approved_payment:
                flash('登录成功！', 'success')
                return redirect(url_for('members'))
            else:
                flash('登录成功！请完成支付验证', 'info')
                return redirect(url_for('payment_manual'))
        else:
            return render_template('login.html', error="用户名或密码错误")
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """用户退出"""
    session.clear()
    flash('已退出登录', 'info')
    return redirect(url_for('index'))

# ============ 密码管理路由 ============

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """修改密码（需要登录）"""
    if request.method == 'POST':
        # 确保从表单获得字符串，避免 None 传入 check_password_hash
        current_password = (request.form.get('current_password') or '').strip()
        new_password = (request.form.get('new_password') or '').strip()
        confirm_password = (request.form.get('confirm_password') or '').strip()
        
        # 获取当前用户
        user = User.query.get(session['user_id'])
        if not user:
            flash('用户不存在或已被删除，请重新登录', 'error')
            session.clear()
            return redirect(url_for('login'))
        
        # 验证当前密码（确保传入的都是 str）
        if not current_password or not check_password_hash(user.password, current_password):
            flash('当前密码错误', 'error')
            return render_template('change_password.html')
        
        # 验证新密码
        if new_password != confirm_password:
            flash('新密码与确认密码不一致', 'error')
            return render_template('change_password.html')
        
        if len(new_password or '') < 6:
            flash('密码长度至少6位', 'error')
            return render_template('change_password.html')
        
        try:
            # 更新密码
            user.password = generate_password_hash(new_password)
            db.session.commit()
            flash('密码修改成功', 'success')
            return redirect(url_for('members'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'密码修改失败: {str(e)}', 'error')
    
    return render_template('change_password.html')

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    """忘记密码 - 请求重置"""
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        
        if not email:
            flash('请输入注册邮箱', 'error')
            return render_template('forgot_password.html')
        
        # 查找用户
        user = User.query.filter_by(email=email).first()
        
        if user:
            try:
                # 生成重置令牌
                import secrets
                token = secrets.token_urlsafe(32)
                expires_at = datetime.now() + timedelta(hours=1)  # 1小时有效
                
                # 删除用户之前的重置令牌
                PasswordReset.query.filter_by(user_id=user.id).delete()
                
                # 创建新的重置令牌
                reset_request = PasswordReset(
                    user_id=user.id,
                    token=token,
                    expires_at=expires_at
                )
                db.session.add(reset_request)
                db.session.commit()
                
                # 生成重置链接（生产环境应发送邮件）
                reset_url = url_for('reset_password', token=token, _external=True)
                
                # 暂时在控制台输出（生产环境应发送邮件）
                print(f"🔐 密码重置链接（用户: {user.email}）:")
                print(f"📧 {reset_url}")
                print(f"⏰ 有效期至: {expires_at.strftime('%Y-%m-%d %H:%M')}")
                
                flash('密码重置链接已生成（请在控制台查看）', 'success')
                
            except Exception as e:
                db.session.rollback()
                flash(f'重置请求失败: {str(e)}', 'error')
        else:
            # 即使邮箱不存在也显示成功，防止邮箱探测
            flash('如果该邮箱已注册，重置链接将发送到您的邮箱', 'info')
        
        return redirect(url_for('forgot_password'))
    
    return render_template('forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    """通过令牌重置密码"""
    # 验证令牌
    reset_request = PasswordReset.query.filter_by(token=token).first()
    
    if not reset_request:
        flash('重置链接无效或已过期', 'error')
        return redirect(url_for('forgot_password'))
    
    if not reset_request.is_valid():
        flash('重置链接已过期', 'error')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        # 确保从表单获得字符串，避免 None 传入 generate_password_hash
        new_password = (request.form.get('new_password') or '').strip()
        confirm_password = (request.form.get('confirm_password') or '').strip()
        
        # 验证密码
        if new_password != confirm_password:
            flash('密码与确认密码不一致', 'error')
            return render_template('reset_password.html', token=token)
        
        if len(new_password) < 6:
            flash('密码长度至少6位', 'error')
            return render_template('reset_password.html', token=token)
        
        try:
            # 更新用户密码
            user = User.query.get(reset_request.user_id)
            user.password = generate_password_hash(new_password)
            
            # 标记令牌为已使用
            reset_request.used = True
            db.session.commit()
            
            flash('密码重置成功，请使用新密码登录', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'密码重置失败: {str(e)}', 'error')
    
    return render_template('reset_password.html', token=token)

# ============ 支付相关路由 ============
@app.route('/payment-manual')
@login_required
def payment_manual():
    """手动支付页面"""
    return render_template('payment_manual.html', 
                          price=MEMBERSHIP_PRICE,
                          admin_email=ADMIN_EMAIL)

@app.route('/submit-payment-proof', methods=['POST'])
@login_required
def submit_payment_proof():
    """提交支付凭证"""
    payment_proof = request.form.get('payment_proof', '').strip()
    
    if not payment_proof:
        flash('请提供支付凭证', 'warning')
        return redirect(url_for('payment_manual'))
    
    try:
        new_payment = Payment(
            user_id=session['user_id'],
            amount=MEMBERSHIP_PRICE,
            payment_method='manual',
            status='pending',
            notes=f"支付凭证: {payment_proof}"
        )
        
        db.session.add(new_payment)
        db.session.commit()
        
        flash('支付凭证已提交，请等待管理员审核', 'success')
        return redirect(url_for('check_payment_status'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'提交失败: {str(e)}', 'error')
        return redirect(url_for('payment_manual'))

@app.route('/check-payment-status')
@login_required
def check_payment_status():
    """检查支付状态"""
    payments = Payment.query.filter_by(user_id=session['user_id'])\
                           .order_by(Payment.create_time.desc()).all()
    has_approved = any(p.status == 'approved' for p in payments)
    
    return render_template('payment_status.html', 
                         payments=payments,
                         has_approved_payment=has_approved,
                         price=MEMBERSHIP_PRICE)

# ============ 会员内容路由 ============
@app.route('/members')
@payment_required
def members():
    """会员专属内容页面"""
    return render_template('members.html')

# ============ 管理员路由 ============
@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    """管理员登录 - 安全版本"""
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if not all([username, password]):
            flash('请填写用户名和密码', 'error')
            return render_template('admin_login.html')
        
        # 从数据库验证管理员
        admin_user = User.query.filter_by(username=username, is_admin=True).first()
        
        if admin_user and check_password_hash(admin_user.password, password):
            session['admin_logged_in'] = True
            session['admin_username'] = username
            flash('管理员登录成功！', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('管理员账号或密码错误', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    """管理员仪表板"""
    stats = {
        'unanswered_count': Question.query.filter_by(answered=False).count(),
        'total_payments': Payment.query.count(),
        'pending_payments': Payment.query.filter_by(status='pending').count(),
        'total_users': User.query.count()
    }
    
    return render_template('admin_dashboard.html', **stats) 

# ============ 专员问答路由 ============

@app.route('/submit-question', methods=['POST'])
@payment_required
def submit_question():
    disclaimer = """
    <div class="alert alert-info">
        <strong>注意：</strong>专员回答仅为个人经验分享，
        不构成专业建议，请谨慎参考。
    </div>
    """
    """用户提交问题 - 修复版本"""
    try:
        # 同时支持表单数据和JSON数据
        if request.is_json:
            data = request.get_json()
            content = data.get('content', '').strip()
        else:
            content = request.form.get('content', '').strip()
        
        print(f"📝 收到问题提交: {content[:100]}...")  # 调试日志
        
        if not content:
            return jsonify({'success': False, 'message': '问题内容不能为空'})
        
        if len(content) < 5:
            return jsonify({'success': False, 'message': '问题内容太短，请详细描述'})
        
        new_question = Question(
            user_id=session['user_id'],
            content=content
        )
        
        db.session.add(new_question)
        db.session.commit()
        
        print(f"✅ 问题提交成功，ID: {new_question.id}")  # 调试日志
        return jsonify({'success': True, 'message': '问题提交成功！有经验人士将在24小时内分享经验'})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ 问题提交失败: {str(e)}")  # 错误日志
        return jsonify({'success': False, 'message': f'提交失败: {str(e)}'})

@app.route('/get-my-questions')
@payment_required
def get_my_questions():
    """获取用户自己的问题列表"""
    questions = Question.query.filter_by(user_id=session['user_id'])\
                             .order_by(Question.create_time.desc()).all()
    
    questions_data = []
    for q in questions:
        questions_data.append({
            'id': q.id,
            'content': q.content,
            'answer': q.answer,
            'answered': q.answered,
            'create_time': q.create_time.strftime('%Y-%m-%d %H:%M'),
            'answer_time': q.answer_time.strftime('%Y-%m-%d %H:%M') if q.answer_time else None
        })
    
    return jsonify({'success': True, 'questions': questions_data})

@app.route('/admin/answer-question/<int:question_id>', methods=['POST'])
@admin_required
def answer_question(question_id):
    """管理员回答问题 - 修复版本"""
    try:
        question = Question.query.get_or_404(question_id)
        
        # 检查请求数据
        if not request.is_json:
            return jsonify({'success': False, 'message': '请求必须是JSON格式'})
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '无效的JSON数据'})
        
        answer_content = data.get('answer', '').strip()
        if not answer_content:
            return jsonify({'success': False, 'message': '回答内容不能为空'})
        
        # 更新问题
        question.answer = answer_content
        question.answered = True
        question.answer_time = datetime.now()
        
        db.session.commit()
        
        print(f"管理员已回答问题 ID: {question_id}")  # 调试日志
        
        return jsonify({'success': True, 'message': '回答提交成功'})
        
    except Exception as e:
        db.session.rollback()
        print(f"回答问题错误: {str(e)}")  # 调试日志
        return jsonify({'success': False, 'message': f'回答失败: {str(e)}'})

@app.route('/admin/payments')
@admin_required
def admin_payments():
    """支付管理 - 优化版本"""
    # 使用优化后的查询方法
    payments = db.session.query(Payment).order_by(Payment.create_time.desc())\
                           .options(db.joinedload(Payment.user))\
                           .all()
    return render_template('admin_payments.html', payments=payments)

@app.route('/admin/update-payment/<int:payment_id>', methods=['POST'])
@admin_required
def update_payment_status(payment_id):
    """更新支付状态"""
    payment = Payment.query.get_or_404(payment_id)
    new_status = request.form.get('status', '')
    
    payment.status = new_status
    payment.process_time = datetime.now()
    db.session.commit()
    
    flash(f'支付状态已更新为 {new_status}', 'success')
    return redirect(url_for('admin_payments'))

@app.route('/admin/users')
@admin_required
def admin_users():
    """用户管理"""
    try:
        users = User.query.order_by(User.create_time.desc()).all()
        return render_template('admin_users.html', 
                             users=users, 
                             now=datetime.now(), 
                             timedelta=timedelta)
    except Exception as e:
        flash(f'用户管理页面加载失败: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete-user/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    """删除用户 - 修复会话版本"""
    try:
        print(f"🔍 开始删除用户 {user_id}")
        
        # 调试：检查会话状态
        print(f"🔍 会话状态检查:")
        print(f"  - session.get('user_id'): {session.get('user_id')}")
        print(f"  - session.get('admin_logged_in'): {session.get('admin_logged_in')}")
        print(f"  - session.get('admin_username'): {session.get('admin_username')}")
        
        # 获取当前管理员ID - 修复版本
        admin_id = session.get('user_id')
        if not admin_id:
            # 如果 user_id 不存在，尝试通过管理员用户名查找
            admin_username = session.get('admin_username')
            if admin_username:
                admin_user = User.query.filter_by(username=admin_username, is_admin=True).first()
                if admin_user:
                    admin_id = admin_user.id
                    print(f"🔍 通过用户名找到管理员ID: {admin_id}")
        
        if not admin_id:
            return jsonify({'success': False, 'message': '管理员会话无效，请重新登录'})
        
        # 防止删除自己
        if user_id == admin_id:
            return jsonify({'success': False, 'message': '不能删除自己的账户'})
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': '用户不存在'})
        
        # 防止删除最后一个管理员
        if user.is_admin:
            admin_count = User.query.filter_by(is_admin=True).count()
            if admin_count <= 1:
                return jsonify({'success': False, 'message': '不能删除最后一个管理员'})
        
        # 删除关联数据
        Payment.query.filter_by(user_id=user_id).delete()
        Question.query.filter_by(user_id=user_id).delete()
        
        # 删除用户
        db.session.delete(user)
        
        # 修复：使用正确的admin_id
        log = AdminLog(
            admin_id=admin_id,  # 使用修复后的admin_id
            action=f'删除用户: {user.username} (ID: {user_id})',
            target_type='user',
            target_id=user_id,
            ip_address=request.remote_addr or 'unknown',
            user_agent=request.headers.get('User-Agent', 'unknown')
        )
        db.session.add(log)
        
        db.session.commit()
        
        print(f"✅ 用户 {user_id} 删除成功")
        return jsonify({'success': True, 'message': '用户已删除'})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ 删除异常: {str(e)}")
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})

# ============ 初始化应用 ============
def init_db():
    """初始化数据库 - PostgreSQL专用版本"""
    with app.app_context():
        try:
            print("🔄 开始初始化PostgreSQL数据库...")
            
            # 创建所有表
            db.create_all()
            print("✅ 数据库表创建完成")

            # 检查AdminLog表是否存在
            from sqlalchemy import inspect
            inspector = inspect(db.engine)
            tables = inspector.get_table_names()
            print(f"📊 数据库中的表: {tables}")
            
            # 创建默认管理员账户
            admin_user = User.query.filter_by(username='huang').first()
            if not admin_user:
                admin_user = User(
                    username='huang',
                    email='942521233@qq.com',
                    password=generate_password_hash('112588'),
                    is_admin=True
                )
                db.session.add(admin_user)
                print("✅ 默认管理员账户已创建: huang / 112588")
            else:
                print("✅ 管理员账户已存在")
            
            # 创建测试用户（方便测试）
            test_user = User.query.filter_by(username='testuser').first()
            if not test_user:
                test_user = User(
                    username='testuser',
                    email='test@example.com',
                    password=generate_password_hash('test123'),
                    is_admin=False
                )
                db.session.add(test_user)
                print("✅ 测试用户已创建: testuser / test123")
            
            db.session.commit()
            print("🎉 PostgreSQL数据库初始化完成！")
            print("💾 用户数据现在将永久保存，不再因部署而丢失！")
            
        except Exception as e:
            db.session.rollback()
            print(f"❌ 数据库初始化错误: {str(e)}")
            
            # 如果是连接错误，提供具体建议
            if "connection" in str(e).lower():
                print("🔧 请检查PostgreSQL连接字符串和网络连接")
            elif "already exists" in str(e).lower():
                print("⚠️  表已存在，可以忽略此错误")
            else:
                raise

@app.route('/admin/make-admin/<int:user_id>', methods=['POST'])
@admin_required
def make_admin(user_id):
    """将用户设为管理员"""
    try:
        user = User.query.get_or_404(user_id)
        
        # 检查是否已经是管理员
        if user.is_admin:
            return jsonify({'success': False, 'message': '用户已经是管理员'})
        
        user.is_admin = True
        db.session.commit()
        
        return jsonify({'success': True, 'message': '用户已设为管理员'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'操作失败: {str(e)}'})

@app.route('/admin/questions')
@admin_required
def admin_questions():
    """问题管理 - 修复版本"""
    try:
        # 使用更明确的查询，确保正确加载用户关系
        questions = db.session.query(Question)\
                             .options(db.joinedload(Question.user))\
                             .order_by(Question.create_time.desc())\
                             .all()
        
        print(f"管理员查看问题: 找到 {len(questions)} 个问题")  # 调试日志
        
        return render_template('admin_questions.html', questions=questions)
        
    except Exception as e:
        print(f"管理员问题查询错误: {str(e)}")  # 调试日志
        flash(f'加载问题列表失败: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/logout')
def admin_logout():
    """管理员退出"""
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    flash('已退出管理员账号', 'info')
    return redirect(url_for('admin_login'))

@app.route('/terms')
def terms():
    """服务条款页面"""
    return render_template('terms.html', ADMIN_EMAIL=ADMIN_EMAIL)

@app.route('/privacy')
def privacy():
    """隐私政策页面"""
    return render_template('privacy.html', ADMIN_EMAIL=ADMIN_EMAIL) 

@app.route('/knowledge-base')
@payment_required
def knowledge_base():
    """知识库页面"""
    return render_template('knowledge_base.html', 
                         content_categories=CONTENT_CATEGORIES,
                         content_modules=CONTENT_MODULES)

# ============ 债务计算器API ============
@app.route('/api/calculate-debt', methods=['POST'])
@login_required
def calculate_debt():
    """债务计算器API"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '无效的请求数据'})
        
        total_debt = float(data.get('total_debt', 0))
        monthly_payment = float(data.get('monthly_payment', 0))
        interest_rate = float(data.get('interest_rate', 12))
        
        if total_debt <= 0 or monthly_payment <= 0:
            return jsonify({'success': False, 'message': '请输入有效的债务金额和月还款额'})
        
        # 计算还款计划
        monthly_rate = interest_rate / 100 / 12
        remaining_debt = total_debt
        months = 0
        total_interest = 0
        payment_plan = []
        
        # 计算还款月数
        while remaining_debt > 0 and months < 600:  # 限制最多50年
            interest = remaining_debt * monthly_rate
            principal = monthly_payment - interest
            
            if principal <= 0:
                return jsonify({
                    'success': False, 
                    'message': '月还款额不足以支付利息，请增加月还款额'
                })
            
            remaining_debt -= principal
            total_interest += interest
            months += 1
            
            # 记录每月还款详情
            payment_plan.append({
                'month': months,
                'principal': round(principal, 2),
                'interest': round(interest, 2),
                'remaining': round(max(remaining_debt, 0), 2)
            })
            
            if months >= 600:
                break
        
        years = months // 12
        remaining_months = months % 12
        
        # 生成建议
        advice = generate_debt_advice(total_debt, monthly_payment, months)
        
        return jsonify({
            'success': True,
            'result': {
                'total_debt': total_debt,
                'monthly_payment': monthly_payment,
                'total_months': months,
                'years': years,
                'remaining_months': remaining_months,
                'total_interest': round(total_interest, 2),
                'total_payment': round(total_debt + total_interest, 2),
                'advice': advice,
                'payment_plan': payment_plan[:12]  # 只返回前12个月的详细计划
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'计算失败: {str(e)}'})

def generate_debt_advice(total_debt, monthly_payment, months):
    """生成债务建议"""
    if months <= 12:
        return {
            'level': 'success',
            'title': '计算示例，仅供参考',
            'content': '此为模拟计算，实际请咨询金融机构。',
            'suggestions': [
                '建议咨询正规金融机构',
                '计算结果仅供参考',
                '请以实际合同为准'
            ]
        }
    elif months <= 36:
        return {
            'level': 'warning',
            'title': '计算示例，仅供参考',
            'content': '此为模拟计算，实际请咨询金融机构。',
            'suggestions': [
                '寻找兼职或副业增加收入',
                '建议优化日常开支',
                '与债权人协商降低利率'
            ]
        }
    else:
        return {
            'level': 'danger',
            'title': '计算示例，仅供参考',
            'content': '此为模拟计算，实际请咨询金融机构。',
            'suggestions': [
                '建议与所有债权人协商还款方案',
                '寻求专业债务咨询服务',
                '制定严格的预算计划',
                '请以实际合同为准'
            ]
        }

# ============ 获取用户进度 ============
@app.route('/api/user-progress')
@payment_required
def get_user_progress():
    """获取用户学习进度"""
    try:
        user_id = session['user_id']
        
        # 获取用户的学习数据（这里需要根据实际数据结构调整）
        completed_courses = 15  # 模拟数据
        completed_steps = 6     # 模拟数据
        in_progress_steps = 3   # 模拟数据
        
        # 计算总体进度
        total_progress = min(100, int((completed_steps / (completed_steps + in_progress_steps)) * 100))
        
        return jsonify({
            'success': True,
            'progress': {
                'total_progress': total_progress,
                'completed_courses': completed_courses,
                'completed_steps': completed_steps,
                'in_progress_steps': in_progress_steps,
                'pending_tasks': 2
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取进度失败: {str(e)}'})

# ============ 工具箱内容API ============
@app.route('/api/tool-content/<tool_type>')
@payment_required
def get_tool_content(tool_type):
    """获取工具箱内容"""
    tools = {
        'harassment': {
            'title': '催收应对经验分享',
            'content': """
                <h4>合法应对催收电话经验分享</h4>
                <ul>
                    <li><strong>保持冷静：</strong>不要与催收人员争吵的经验分享</li>
                    <li><strong>录音取证：</strong>所有通话都要录音保存的经验</li>
                    <li><strong>明确表达：</strong>表明还款意愿但暂时困难的经验</li>
                    <li><strong>了解权利：</strong>催收不得骚扰家人朋友的知识</li>
                    <li><strong>投诉渠道：</strong>遭遇违规催收可拨打12378投诉的经验</li>
                </ul>
                <div class="alert alert-warning mt-3">
                    <strong>注意：</strong>如果催收人员威胁、辱骂或上门骚扰，立即向银保监会投诉的经验分享。
                </div>
            """
        },
        'legal': {
            'title': '法律知识分享',
            'content': """
                <h4>权益保护知识分享</h4>
                <ul>
                    <li><strong>个人信息权：</strong>催收不得泄露债务信息的经验</li>
                    <li><strong>休息权：</strong>晚上10点至早上8点不得催收的经验</li>
                    <li><strong>名誉权：</strong>不得公开侮辱、诽谤的经验</li>
                    <li><strong>协商权：</strong>有权要求协商还款方案的经验</li>
                </ul>
                <h4 class="mt-4">常见违法行为识别经验</h4>
                <ul>
                    <li>爆通讯录、联系无关第三人的识别</li>
                    <li>P图、发假律师函的识别</li>
                    <li>上门骚扰、威胁的应对经验</li>
                    <li>冒充公检法人员的识别</li>
                </ul>
                <div class="alert alert-info mt-3">
                    <strong>维权方式经验分享：</strong>收集证据 → 向银保监会12378投诉 → 必要时报警的经验
                </div>
            """
        },
        'psychological': {
            'title': '心理调适经验分享',
            'content': """
                <h4>缓解债务焦虑经验分享</h4>
                <ul>
                    <li><strong>接受现实：</strong>债务是暂时困难，不是人生终点的经验</li>
                    <li><strong>分解目标：</strong>将大目标分解为可执行的小步骤的经验</li>
                    <li><strong>寻求支持：</strong>与家人沟通或加入支持群体的经验</li>
                    <li><strong>保持运动：</strong>每天30分钟运动缓解压力的经验</li>
                    <li><strong>正面思考：</strong>关注解决方案而非问题本身的经验</li>
                </ul>
                <h4 class="mt-4">紧急心理支持资源</h4>
                <p>如果感到极度焦虑、抑郁或有自杀念头，请立即寻求专业帮助：</p>
                <ul>
                    <li>心理援助热线：12320</li>
                    <li>危机干预热线：800-810-1117</li>
                    <li>当地心理卫生中心</li>
                </ul>
            """
        }
    }
    
    tool = tools.get(tool_type)
    if tool:
        return jsonify({'success': True, 'tool': tool})
    else:
        return jsonify({'success': False, 'message': '工具不存在'})

# ============ 优化：问答功能 ============
@app.route('/api/submit-question', methods=['POST'])
@payment_required
def api_submit_question():
    """API版本的问题提交"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': '无效的请求数据'})
    
    content = data.get('content', '').strip()
    
    if not content:
        return jsonify({'success': False, 'message': '问题内容不能为空'})
    
    try:
        new_question = Question(
            user_id=session['user_id'],
            content=content
        )
        
        db.session.add(new_question)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '问题提交成功！有经验人士将在24小时内分享经验'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'提交失败: {str(e)}'})

@app.route('/api/my-questions')
@payment_required
def api_my_questions():
    """API版本的用户问题列表"""
    try:
        questions = Question.query.filter_by(user_id=session['user_id'])\
                                 .order_by(Question.create_time.desc()).all()
        
        questions_data = []
        for q in questions:
            questions_data.append({
                'id': q.id,
                'content': q.content,
                'answer': q.answer,
                'answered': q.answered,
                'create_time': q.create_time.strftime('%Y-%m-%d %H:%M'),
                'answer_time': q.answer_time.strftime('%Y-%m-%d %H:%M') if q.answer_time else None
            })
        
        return jsonify({'success': True, 'questions': questions_data})
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取问题列表失败: {str(e)}'})

# ============ 资源下载 ============
@app.route('/download/<resource_type>')
@payment_required
def download_resource(resource_type):
    """资源下载"""
    resources = {
        'debt-template': {
            'filename': '债务管理经验模板.xlsx',
            'description': '债务管理经验模板分享'
        },
        'negotiation-guide': {
            'filename': '协商经验分享指南.pdf',
            'description': '协商经验分享指南'
        },
        'legal-rights': {
            'filename': '法律知识分享手册.pdf',
            'description': '债务相关法律知识分享'
        }
    }
    
    resource = resources.get(resource_type)
    if resource:
        # 这里应该返回实际的文件
        # 暂时返回成功消息
        flash(f'开始下载: {resource["description"]}', 'success')
        return jsonify({'success': True, 'message': f'开始下载 {resource["description"]}'})
    else:
        return jsonify({'success': False, 'message': '资源不存在'})
    
    # ============ 心理建设支持路由 ============
@app.route('/psychological-support')
@payment_required
def psychological_support():
    """心理建设支持页面 - 修复版本"""
    try:
        print("🔄 正在渲染心理建设支持页面...")
        user_id = session.get('user_id')
        username = session.get('username')
        print(f"👤 用户: {username} (ID: {user_id})")
        
        return render_template('psychological-support.html')
    except Exception as e:
        print(f"❌ 心理建设支持页面错误: {e}")
        flash(f'页面加载失败: {str(e)}', 'error')
        return redirect(url_for('members'))
    
# ============ 重定向路由（兼容旧链接） ============
@app.route('/members.html')
@payment_required
def redirect_members():
    """将 members.html 重定向到 /members"""
    return redirect(url_for('members'))

@app.route('/psychological-support.html')
@payment_required
def redirect_psychological_support():
    """将 psychological-support.html 重定向到 /psychological-support"""
    return redirect(url_for('psychological_support'))

@app.route('/debt-management-course.html')
@payment_required
def redirect_debt_management_course():
    """重定向债务管理课程"""
    return redirect(url_for('debt_management_course'))

@app.route('/negotiation-guide.html')
@payment_required
def redirect_negotiation_guide():
    """重定向协商话术指南"""
    return redirect(url_for('negotiation_guide'))

@app.route('/psychological-course')
@payment_required
def psychological_course():
    """心理调适经验分享页面"""
    try:
        return render_template('psychological-course.html')
    except Exception as e:
        print(f"心理调适经验分享页面错误: {e}")
        return "心理调适经验分享页面暂时不可用", 500
    
@app.route('/income-projects')
@payment_required
def income_projects():
    """创收项目经验分享页面"""
    try:
        return render_template('income-projects.html')
    except Exception as e:
        print(f"创收项目经验分享页面错误: {e}")
        return "创收项目经验分享页面暂时不可用", 500
    
# ============ 调试路由 ============
@app.route('/debug/questions')
@admin_required
def debug_questions():
    """调试问题数据"""
    try:
        # 检查所有问题
        all_questions = Question.query.all()
        print(f"总问题数量: {len(all_questions)}")
        
        # 检查未回答问题
        unanswered = Question.query.filter_by(answered=False).all()
        print(f"未回答问题数量: {len(unanswered)}")
        
        # 检查用户关联
        for q in all_questions:
            user = User.query.get(q.user_id)
            print(f"问题ID: {q.id}, 用户ID: {q.user_id}, 用户名: {user.username if user else '用户不存在'}, 已回答: {q.answered}")
        
        return jsonify({
            'total_questions': len(all_questions),
            'unanswered_questions': len(unanswered),
            'questions': [
                {
                    'id': q.id,
                    'user_id': q.user_id,
                    'username': User.query.get(q.user_id).username if User.query.get(q.user_id) else 'Unknown',
                    'content': q.content[:50] + '...' if len(q.content) > 50 else q.content,
                    'answered': q.answered,
                    'create_time': q.create_time.strftime('%Y-%m-%d %H:%M')
                } for q in all_questions
            ]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

# ============ 债务管理相关路由 ============

@app.route('/api/debt-management/progress')
@payment_required
def get_debt_management_progress():
    """获取债务管理学习进度"""
    try:
        user_id = session['user_id']
        
        # 这里可以从数据库获取用户的实际进度
        # 暂时返回模拟数据
        return jsonify({
            'success': True,
            'progress': {
                'total_progress': 65,
                'completed_steps': 6,
                'in_progress_steps': 3,
                'total_steps': 10,
                'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M')
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取进度失败: {str(e)}'})

@app.route('/api/debt-management/update-progress', methods=['POST'])
@payment_required
def update_debt_management_progress():
    """更新学习进度"""
    try:
        data = request.get_json()
        step_completed = data.get('step')
        
        # 这里可以更新数据库中的用户进度
        # 暂时返回成功响应
        return jsonify({
            'success': True,
            'message': f'步骤 {step_completed} 已完成',
            'progress': 65  # 模拟进度
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'更新进度失败: {str(e)}'})

@app.route('/debt-management-course')
@payment_required
def debt_management_course():
    """债务管理经验分享页面 - 完整版本"""
    try:
        course_content = {
            'title': '债务管理经验分享',
            'sections': [
                {
                    'title': '停止以贷养贷经验',
                    'icon': 'ban',
                    'content': """
                        <h4>🛑 为什么必须停止以贷养贷？</h4>
                        <div class="alert alert-danger">
                            <strong>恶性循环警告：</strong>以贷养贷就像在流沙中挣扎，越挣扎陷得越深！
                        </div>
                        
                        <h5>💸 以贷养贷的真实代价经验分享</h5>
                        <div class="row">
                            <div class="col-md-6">
                                <div class="card border-danger mb-3">
                                    <div class="card-body">
                                        <h6 class="card-title text-danger"><i class="fas fa-chart-line me-2"></i>利息翻倍经验</h6>
                                        <p class="card-text">新贷款利息 + 旧债务利息 = 双重利息负担的经验</p>
                                        <small class="text-muted">例：5万债务一年可能多付1-2万利息的经验分享</small>
                                    </div>
                                </div>
                            </div>
                            <div class="col-md-6">
                                <div class="card border-danger mb-3">
                                    <div class="card-body">
                                        <h6 class="card-title text-danger"><i class="fas fa-snowflake me-2"></i>债务雪球经验</h6>
                                        <p class="card-text">小额债务滚成大额债务，最终无法控制的经验</p>
                                        <small class="text-muted">很多大额负债都是从几千元开始的经验分享</small>
                                    </div>
                                </div>
                            </div>
                        </div>

                        <h5>🚫 立即停止的实战经验分享</h5>
                        <div class="table-responsive">
                            <table class="table table-striped">
                                <thead>
                                    <tr>
                                        <th>行动</th>
                                        <th>具体做法经验</th>
                                        <th>效果</th>
                                    </tr>
                                </thead>
                                <tbody>
                                    <tr>
                                        <td><strong>注销信用卡</strong></td>
                                        <td>剪掉所有信用卡，消除透支可能的经验</td>
                                        <td>立即切断透支渠道</td>
                                    </tr>
                                    <tr>
                                        <td><strong>删除借贷APP</strong></td>
                                        <td>卸载所有网贷应用程序的经验</td>
                                        <td>消除冲动借贷可能</td>
                                    </tr>
                                </tbody>
                            </table>
                        </div>
                        
                        <div class="alert alert-success mt-4">
                            <h6><i class="fas fa-lightbulb me-2"></i>经验分享</h6>
                            <p class="mb-0">"我曾经欠款30万，通过停止以贷养贷，制定科学还款计划，3年时间成功上岸。现在回想，停止养贷是我做过最正确的决定！" — 经验分享</p>
                        </div>
                    """,
                    'tools': [
                        {
                            'name': '债务计算器',
                            'icon': 'calculator',
                            'color': 'primary',
                            'description': '计算您的真实债务成本和还款周期',
                            'button_text': '使用计算器',
                            'button_icon': 'calculator',
                            'action': 'window.location.href="/members#debtCalculator"'
                        }
                    ],
                    'actions': [
                        {'text': '我已停止以贷养贷', 'type': 'success', 'step': 'stop_borrowing'},
                        {'text': '需要更多帮助', 'type': 'warning', 'step': 'need_help_stop'}
                    ]
                }
            ]
        }
        
        return render_template('debt_management_course.html', 
                             course=course_content,
                             progress=65)
    except Exception as e:
        print(f"债务管理经验分享页面错误: {e}")
        return "债务管理经验分享页面暂时不可用", 500
    
@app.route('/download/debt-management-template')
@payment_required
def download_debt_template():
    """下载债务管理经验模板"""
    try:
        print("🔍 DEBUG: 下载债务管理经验模板")
        
        # 尝试创建专业的Excel文件
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            
            # ===== 债务清单表 =====
            ws_debts = wb.active
            ws_debts.title = "债务清单参考表"
            
            # 设置表头样式
            headers = ['序号', '债权人', '债务类型', '总借款金额(元)', '已还金额(元)', 
                      '剩余本金(元)', '年利率(%)', '每月最低还款', '逾期状态', 
                      '最后还款日', '紧急程度', '还款优先级', '备注']
            
            for col, header in enumerate(headers, 1):
                cell = ws_debts.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 添加示例数据和公式
            example_data = [
                [1, '招商银行信用卡', '信用卡', 50000, 5000, 45000, 18.25, 2500, '逾期', '2024-03-15', '紧急', 1, '经验分享：已协商分期60期'],
                [2, '支付宝借呗', '网贷', 30000, 0, 30000, 15.5, 1800, '正常', '2024-03-20', '高息', 2, '经验分享：正常还款中'],
                [3, '微信微粒贷', '网贷', 20000, 2000, 18000, 16.8, 1200, '逾期', '2024-03-10', '紧急', 3, '经验分享：催收中，需协商'],
            ]
            
            for row, data in enumerate(example_data, 2):
                for col, value in enumerate(data, 1):
                    cell = ws_debts.cell(row=row, column=col, value=value)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                       top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 添加汇总行
            summary_row = len(example_data) + 3
            ws_debts.cell(row=summary_row, column=4, value="总债务金额:").font = Font(bold=True)
            ws_debts.cell(row=summary_row, column=5, value="=SUM(D2:D4)").font = Font(bold=True, color="FF0000")
            
            ws_debts.cell(row=summary_row+1, column=4, value="剩余债务总额:").font = Font(bold=True)
            ws_debts.cell(row=summary_row+1, column=5, value="=SUM(F2:F4)").font = Font(bold=True, color="FF0000")
            
            ws_debts.cell(row=summary_row+2, column=4, value="月最低还款总额:").font = Font(bold=True)
            ws_debts.cell(row=summary_row+2, column=5, value="=SUM(H2:H4)").font = Font(bold=True, color="FF0000")
            
            # 设置列宽
            column_widths = [8, 15, 12, 15, 15, 15, 12, 15, 12, 15, 12, 12, 20]
            for col, width in enumerate(column_widths, 1):
                ws_debts.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # ===== 还款计划表 =====
            ws_plan = wb.create_sheet("还款计划参考表")
            
            plan_headers = ['月份', '总收入', '必要支出', '可还款金额', '信用卡还款', '网贷还款', 
                          '其他还款', '应急储备', '剩余债务', '完成情况', '备注']
            
            for col, header in enumerate(plan_headers, 1):
                cell = ws_plan.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # ===== 使用指南表 =====
            ws_guide = wb.create_sheet("经验分享指南")
            guide_content = [
                ["债务管理经验分享模板使用指南"],
                [""],
                ["📋 债务清单参考表使用说明"],
                ["1. 填写所有债务的详细信息（经验分享）"],
                ["   - 如实填写每个债权人的债务情况"],
                ["   - 准确计算剩余本金和利息"],
                ["   - 根据紧急程度和利率确定还款优先级"],
                [""],
                ["📅 还款计划参考表使用说明"],
                ["1. 根据收入制定月度还款计划（经验分享）"],
                ["   - 优先处理紧急和高息债务"],
                ["   - 确保还款金额在承受范围内"],
                ["   - 保留必要的应急资金"],
                [""],
                ["💡 经验分享建议"],
                ["1. 每周更新一次表格，跟踪进度"],
                ["2. 不要以贷养贷，切断恶性循环"],
                ["3. 与债权人保持沟通，积极协商"],
                ["4. 坚持执行计划，不要中途放弃"],
                ["5. 寻求专业帮助 if needed"],
                [""],
                ["🎯 还款优先级策略（经验分享）"],
                ["第一优先级：已逾期、可能起诉的债务"],
                ["第二优先级：高利率（>15%）的债务"],
                ["第三优先级：正常还款中的常规债务"],
                ["第四优先级：亲友借款等无息债务"],
                [""],
                ["📞 紧急求助资源"],
                ["银保监会投诉热线：12378"],
                ["心理援助热线：12320"],
                ["上岸翻身营经验分享：会员专属"],
                [""],
                ["生成时间：{}".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))],
                ["上岸翻身营 - 债务管理经验分享平台"]
            ]
            
            for row, content in enumerate(guide_content, 1):
                cell = ws_guide.cell(row=row, column=1, value=content[0])
                if row == 1:
                    cell.font = Font(bold=True, size=14, color="366092")
                elif content[0] and any(marker in content[0] for marker in ["📋", "📅", "💡", "🎯", "📞"]):
                    cell.font = Font(bold=True, color="366092")
            
            # 返回Excel文件
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                as_attachment=True,
                download_name='债务管理经验分享模板.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except ImportError:
            # 如果没有openpyxl，创建专业的CSV文件
            template_content = create_professional_csv_template()
            output = StringIO()
            output.write(template_content)
            output.seek(0)
            
            return send_file(
                BytesIO(output.getvalue().encode('utf-8')),
                as_attachment=True,
                download_name='债务管理经验分享模板.csv',
                mimetype='text/csv'
            )
            
    except Exception as e:
        print(f"❌ 下载错误: {e}")
        flash(f'下载失败: {str(e)}', 'error')
        return redirect(url_for('members'))

def create_professional_csv_template():
    """创建专业的CSV格式债务管理表格"""
    template_content = """上岸翻身营 - 债务管理经验分享模板
生成时间：{}

=== 债务清单参考表 ===
序号,债权人,债务类型,总借款金额(元),已还金额(元),剩余本金(元),年利率(%),每月最低还款,逾期状态,最后还款日,紧急程度,还款优先级,备注
1,招商银行信用卡,信用卡,50000,5000,45000,18.25,2500,逾期,2024-03-15,紧急,1,经验分享：已协商分期60期
2,支付宝借呗,网贷,30000,0,30000,15.5,1800,正常,2024-03-20,高息,2,经验分享：正常还款中
3,微信微粒贷,网贷,20000,2000,18000,16.8,1200,逾期,2024-03-10,紧急,3,经验分享：催收中，需协商

汇总：
总债务金额,100000
剩余债务总额,93000
月最低还款总额,5500

=== 还款计划参考表 ===
月份,总收入,必要支出,可还款金额,信用卡还款,网贷还款,其他还款,应急储备,剩余债务,完成情况,备注
2024-03,8000,4000,4000,2500,1500,0,0,89000,进行中,首月执行经验分享
2024-04,8000,4000,4000,2500,1500,0,0,85000,计划中,坚持计划经验分享

=== 经验分享指南 ===
📋 债务清单参考表使用说明
1. 填写所有债务的详细信息（经验分享）
   - 如实填写每个债权人的债务情况
   - 准确计算剩余本金和利息
   - 根据紧急程度和利率确定还款优先级

📅 还款计划参考表使用说明  
1. 根据收入制定月度还款计划（经验分享）
   - 优先处理紧急和高息债务
   - 确保还款金额在承受范围内
   - 保留必要的应急资金

💡 经验分享建议
1. 每周更新一次表格，跟踪进度
2. 不要以贷养贷，切断恶性循环
3. 与债权人保持沟通，积极协商
4. 坚持执行计划，不要中途放弃
5. 寻求专业帮助 if needed

🎯 还款优先级策略（经验分享）
第一优先级：已逾期、可能起诉的债务
第二优先级：高利率（>15%）的债务  
第三优先级：正常还款中的常规债务
第四优先级：亲友借款等无息债务

📞 紧急求助资源
银保监会投诉热线：12378
心理援助热线：12320
上岸翻身营经验分享：会员专属

上岸翻身营 - 债务管理经验分享平台
为您提供全方位的债务解决方案经验分享
""".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    return template_content

# ============ 学习进度管理 ============

class LearningProgress(db.Model):
    """学习进度模型"""
    __tablename__ = "learning_progress"
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    course_id = db.Column(db.String(50), nullable=False, index=True)
    step_completed = db.Column(db.String(100), nullable=False)
    progress_percentage = db.Column(db.Integer, default=0, nullable=False)
    time_spent = db.Column(db.Integer, default=0)  # 学习时长（分钟）
    completed = db.Column(db.Boolean, default=False, nullable=False)
    last_updated = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    create_time = db.Column(db.DateTime, default=datetime.now)
    
    def __init__(self, user_id: int, course_id: str, step_completed: str, 
                 progress_percentage: int = 0, time_spent: int = 0):
        self.user_id = user_id
        self.course_id = course_id
        self.step_completed = step_completed
        self.progress_percentage = progress_percentage
        self.time_spent = time_spent

# 债务管理课程结构
DEBT_MANAGEMENT_COURSE = {
    'id': 'debt-management',
    'title': '债务管理经验分享',
    'description': '交流债务管理经验，分享制定还款计划的思路',
    'total_steps': 10,
    'estimated_time': '8-12小时',
    'level': '初级到高级',
    'sections': [
        {
            'id': 'foundation',
            'title': '基础认知',
            'steps': [
                {'id': 'stop_borrowing', 'title': '停止以贷养贷经验', 'duration': 30},
                {'id': 'debt_assessment', 'title': '全面评估债务经验', 'duration': 45},
                {'id': 'mindset_change', 'title': '建立正确心态经验', 'duration': 25}
            ]
        },
        {
            'id': 'planning',
            'title': '制定计划经验',
            'steps': [
                {'id': 'debt_inventory', 'title': '制作债务清单经验', 'duration': 60},
                {'id': 'income_analysis', 'title': '分析收支状况经验', 'duration': 45},
                {'id': 'repayment_strategy', 'title': '选择还款策略经验', 'duration': 50}
            ]
        },
        {
            'id': 'implementation',
            'title': '执行优化经验',
            'steps': [
                {'id': 'negotiation_skills', 'title': '债务协商技巧经验', 'duration': 55},
                {'id': 'legal_protection', 'title': '法律权益保护经验', 'duration': 40},
                {'id': 'psychological_support', 'title': '心理支持疏导经验', 'duration': 35},
                {'id': 'long_term_planning', 'title': '长期财务规划经验', 'duration': 50}
            ]
        }
    ]
}

@app.route('/api/learning-progress', methods=['GET'])
@payment_required
def get_learning_progress():
    """获取用户学习进度 - 专业版本"""
    try:
        user_id = session['user_id']
        
        # 获取用户的学习记录
        progress_records = LearningProgress.query.filter_by(
            user_id=user_id, 
            course_id='debt-management'
        ).all()
        
        # 计算总体进度
        completed_steps = set([p.step_completed for p in progress_records])
        total_steps = DEBT_MANAGEMENT_COURSE['total_steps']
        progress_percentage = min(100, int((len(completed_steps) / total_steps) * 100))
        
        # 计算学习时长
        total_time_spent = sum([p.time_spent for p in progress_records])
        
        # 构建详细进度信息
        sections_progress = []
        for section in DEBT_MANAGEMENT_COURSE['sections']:
            section_completed = 0
            section_total = len(section['steps'])
            
            for step in section['steps']:
                if step['id'] in completed_steps:
                    section_completed += 1
            
            sections_progress.append({
                'id': section['id'],
                'title': section['title'],
                'completed': section_completed,
                'total': section_total,
                'progress': int((section_completed / section_total) * 100) if section_total > 0 else 0
            })
        
        # 获取最近学习活动
        recent_activity = []
        for record in sorted(progress_records, key=lambda x: x.last_updated, reverse=True)[:5]:
            # 查找步骤标题
            step_title = "未知步骤"
            for section in DEBT_MANAGEMENT_COURSE['sections']:
                for step in section['steps']:
                    if step['id'] == record.step_completed:
                        step_title = step['title']
                        break
            
            recent_activity.append({
                'step': step_title,
                'completion_time': record.last_updated.strftime('%Y-%m-%d %H:%M'),
                'time_spent': record.time_spent
            })
        
        return jsonify({
            'success': True,
            'progress': {
                'overall_progress': progress_percentage,
                'completed_steps': len(completed_steps),
                'total_steps': total_steps,
                'total_time_spent': total_time_spent,
                'estimated_remaining': max(0, (total_steps - len(completed_steps)) * 45),  # 预估剩余时间
                'started_learning': len(progress_records) > 0,
                'last_activity': progress_records[0].last_updated.strftime('%Y-%m-%d %H:%M') if progress_records else None
            },
            'sections': sections_progress,
            'recent_activity': recent_activity,
            'course_info': {
                'title': DEBT_MANAGEMENT_COURSE['title'],
                'description': DEBT_MANAGEMENT_COURSE['description'],
                'level': DEBT_MANAGEMENT_COURSE['level'],
                'estimated_time': DEBT_MANAGEMENT_COURSE['estimated_time']
            }
        })
        
    except Exception as e:
        print(f"获取学习进度错误: {e}")
        return jsonify({
            'success': False, 
            'message': f'获取学习进度失败: {str(e)}',
            'progress': {
                'overall_progress': 0,
                'completed_steps': 0,
                'total_steps': DEBT_MANAGEMENT_COURSE['total_steps'],
                'total_time_spent': 0,
                'estimated_remaining': DEBT_MANAGEMENT_COURSE['total_steps'] * 45,
                'started_learning': False,
                'last_activity': None
            },
            'sections': [],
            'recent_activity': [],
            'course_info': DEBT_MANAGEMENT_COURSE
        })

@app.route('/api/update-learning-progress', methods=['POST'])
@payment_required
def update_learning_progress():
    """更新学习进度"""
    try:
        data = request.get_json()
        user_id = session['user_id']
        step_id = data.get('step_id')
        time_spent = data.get('time_spent', 0)
        
        if not step_id:
            return jsonify({'success': False, 'message': '步骤ID不能为空'})
        
        # 检查是否已经记录过
        existing_record = LearningProgress.query.filter_by(
            user_id=user_id,
            course_id='debt-management',
            step_completed=step_id
        ).first()
        
        if not existing_record:
            # 创建新的进度记录
            new_progress = LearningProgress(
                user_id=user_id,
                course_id='debt-management',
                step_completed=step_id,
                time_spent=time_spent
            )
            db.session.add(new_progress)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '学习进度已更新'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'更新进度失败: {str(e)}'})

@app.route('/learning-dashboard')
@payment_required
def learning_dashboard():
    """学习经验分享仪表板页面"""
    return render_template('learning_dashboard.html')

# ============ 专业模板内容 ============

@app.route('/api/professional-tools')
@payment_required
def get_professional_tools():
    """获取专业工具列表"""
    tools = [
        {
            'id': 'debt-calculator',
            'name': '债务计算器',
            'description': '计算还款周期、利息成本，分享制定还款方案的思路',
            'icon': 'calculator',
            'color': 'primary',
            'features': ['多债务同时计算', '利息对比分析', '还款方案参考'],
            'button_text': '开始计算',
            'url': '/members#debtCalculator'
        },
        {
            'id': 'debt-template',
            'name': '债务管理经验模板',
            'description': 'Excel经验模板，计算参考、图表分析、进度跟踪',
            'icon': 'file-excel',
            'color': 'success',
            'features': ['债务清单参考', '还款计划表', '进度可视化'],
            'button_text': '下载模板',
            'url': '/download/debt-management-template'
        },
        {
            'id': 'negotiation-guide',
            'name': '债务协商经验分享',
            'description': '协商经验分享、法律知识、应对策略参考',
            'icon': 'comments',
            'color': 'info',
            'features': ['协商流程经验', '法律条款分享', '案例经验参考'],
            'button_text': '查看经验',
            'url': '/knowledge-base'
        },
        {
            'id': 'progress-tracker',
            'name': '进度跟踪经验',
            'description': '可视化学习进度、债务减少趋势、经验分享',
            'icon': 'chart-line',
            'color': 'warning',
            'features': ['学习进度可视化', '债务减少曲线', '经验分享'],
            'button_text': '查看进度',
            'url': '/learning-dashboard'
        }
    ]
    
    return jsonify({'success': True, 'tools': tools})

# ============ 初始化学习进度表 ============
def init_learning_progress():
    """初始化学习进度表"""
    with app.app_context():
        try:
            # 检查表是否存在，如果不存在则创建
            db.create_all()
            print("✅ 学习进度表初始化完成")
        except Exception as e:
            print(f"❌ 学习进度表初始化失败: {e}")

# ============ 协商话术管理 ============

# 协商话术分类和内容
NEGOTIATION_PHRASES = {
    'categories': [
        {
            'id': 'bank',
            'name': '银行协商经验',
            'icon': 'university',
            'color': 'primary',
            'description': '信用卡、银行贷款等银行机构协商经验分享'
        },
        {
            'id': 'online_loan',
            'name': '网贷平台经验',
            'icon': 'mobile-alt',
            'color': 'info',
            'description': '各类网贷平台协商还款经验分享'
        },
        {
            'id': 'legal',
            'name': '法律知识分享',
            'icon': 'balance-scale',
            'color': 'warning',
            'description': '法律法规知识和维权经验分享'
        },
        {
            'id': 'psychological',
            'name': '心理技巧经验',
            'icon': 'brain',
            'color': 'success',
            'description': '沟通心理技巧和情绪管理经验分享'
        }
    ],
    'phrases': [
        # 银行协商话术（已存在的）
        {
            'id': 1,
            'category': 'bank',
            'title': '信用卡逾期协商经验',
            'content': """
尊敬的客服您好，我是贵行信用卡持卡人[姓名]，卡号尾号[XXXX]。由于近期遇到一些经济困难，暂时无法按时全额还款，但我有强烈的还款意愿。希望能与贵行协商一个双方都能接受的还款方案。

我目前的情况是：[简要说明困难原因，如失业、疾病、家庭变故等]。但我有稳定的[收入来源]，每月可以拿出[具体金额]用于还款。

请问是否可以申请：
1. 利息和违约金的减免
2. 分期还款方案（[期数]期）
3. 停止催收骚扰

我愿意提供相关证明材料，希望能得到您的理解和支持。
            """,
            'difficulty': '初级',
            'usage_count': 1250,
            'success_rate': 85,
            'tags': ['开场白', '信用卡', '逾期'],
            'key_points': [
                '态度诚恳，表达还款意愿',
                '说明困难原因但不过多抱怨',
                '提出具体可行的还款方案',
                '主动要求提供证明材料'
            ]
        },
        {
            'id': 2,
            'category': 'bank',
            'title': '个性化分期还款经验',
            'content': """
您好，根据《商业银行信用卡业务监督管理办法》第70条规定，在特殊情况下，确认信用卡欠款金额超出持卡人还款能力、且持卡人仍有还款意愿的，发卡银行可以与持卡人平等协商，达成个性化分期还款协议。

我目前的情况符合上述规定，希望申请个性化分期还款，协议最长可以分期60期。我每月可还款[金额]元，希望贵行能够考虑我的实际情况。

如需要，我可以提供：
- 收入证明
- 困难情况说明
- 征信报告
- 其他相关材料

请告知具体的申请流程和所需材料，我会积极配合。
            """,
            'difficulty': '中级',
            'usage_count': 890,
            'success_rate': 78,
            'tags': ['分期还款', '法律法规', '证明材料'],
            'key_points': [
                '引用具体法规条款',
                '明确分期期数要求',
                '列出可提供的证明材料',
                '表达配合态度'
            ]
        },
        
        # ===== 网贷平台经验分享（新增） =====
        {
            'id': 3,
            'category': 'online_loan',
            'title': '网贷平台延期还款申请经验',
            'content': """
您好，我在贵平台的借款[合同编号]目前因临时困难无法按时还款，但绝非恶意拖欠。希望能申请延期[时间]还款。

我目前的情况：[说明具体困难，如失业、疾病、家庭变故等]。预计在[时间]后情况会好转，届时可以正常还款。

申请事项：
1. 延期至[具体日期]还款
2. 期间停止计算罚息
3. 暂停催收联系

我愿意支付正常的借款利息，只是需要一些时间周转。请考虑我的申请，谢谢！

**经验分享要点：**
1. 主动联系，不要等逾期后再处理
2. 提供证明材料（如失业证明、医院证明等）
3. 表达还款意愿，强调非恶意拖欠
4. 要求合理的延期期限
            """,
            'difficulty': '初级',
            'usage_count': 980,
            'success_rate': 72,
            'tags': ['延期还款', '网贷协商', '暂停催收'],
            'key_points': [
                '主动联系说明情况',
                '提供证明材料',
                '表达非恶意拖欠意愿',
                '要求合理延期期限'
            ]
        },
        {
            'id': 4,
            'category': 'online_loan',
            'title': '网贷高利率协商经验分享',
            'content': """
您好，我注意到贵平台的借款利率较高，综合年化利率达到[利率]%，超过了国家规定的民间借贷利率司法保护上限。

根据最高人民法院的相关规定，借贷利率超过合同成立时一年期LPR四倍的部分不受法律保护。我愿意偿还合法范围内的本息，但超出部分希望能予以减免。

我提议：
- 偿还本金+合法利息（LPR四倍以内）
- 制定可行的还款计划
- 结清后开具结清证明

如果贵平台坚持要求支付超出法律保护范围的利息，我将不得不向金融监管部门投诉维权。

**法律依据参考：**
1. 最高人民法院《关于审理民间借贷案件适用法律若干问题的规定》
2. 借贷利率不得超过合同成立时一年期LPR四倍
3. 超出部分法律不予支持
            """,
            'difficulty': '高级',
            'usage_count': 540,
            'success_rate': 82,
            'tags': ['利率协商', '法律维权', '监管部门投诉'],
            'key_points': [
                '指出利率过高问题',
                '引用法律依据',
                '提出合理还款方案',
                '表明维权决心'
            ]
        },
        {
            'id': 5,
            'category': 'online_loan',
            'title': '网贷平台一次性结清协商经验',
            'content': """
您好，关于我在贵平台的借款[合同编号]，因目前经济困难，无法按期还款，但希望一次性结清债务。

我提议：
1. 减免所有罚息和违约金
2. 只偿还本金和合法利息
3. 给予一定比例的本金减免

我目前的困难是：[说明具体困难]。如果能达成一次性结清协议，我可以向亲友借款或通过其他方式筹集资金。

**经验分享：**
1. 一次性结清通常能争取到更好的减免
2. 准备困难证明材料
3. 强调筹款的困难性
4. 保持耐心，可能需要多次协商
            """,
            'difficulty': '中级',
            'usage_count': 430,
            'success_rate': 65,
            'tags': ['一次性结清', '减免协商', '网贷'],
            'key_points': [
                '提出一次性结清方案',
                '要求减免不合理费用',
                '提供困难证明',
                '保持协商耐心'
            ]
        },
        
        # ===== 法律知识分享（新增） =====
        {
            'id': 6,
            'category': 'legal',
            'title': '违规催收应对法律知识',
            'content': """
根据《中华人民共和国网络安全法》和《商业银行信用卡业务监督管理办法》规定，催收行为必须合法合规。

我目前遭遇到的以下行为涉嫌违规：
1. 非工作时间频繁拨打电话（晚上10点后、早上8点前）
2. 骚扰无关第三人（家人、同事、朋友）
3. 使用威胁、辱骂性语言
4. 冒充司法人员、发送虚假法律文书
5. 泄露债务信息给第三方

**法律知识分享：**
1. 《民法典》第1032条：自然人享有隐私权
2. 《治安管理处罚法》第42条：骚扰他人可处拘留或罚款
3. 银保监会《关于规范商业银行信用卡催收行为的通知》
4. 催收不得使用暴力、威胁、恐吓等手段

**维权步骤：**
1. 录音录像保存证据
2. 向平台客服正式投诉
3. 向银保监会12378投诉
4. 向公安机关报案（如涉及威胁、暴力）
            """,
            'difficulty': '中级',
            'usage_count': 1120,
            'success_rate': 88,
            'tags': ['违规催收', '法律维权', '投诉渠道'],
            'key_points': [
                '列举具体违规行为',
                '引用相关法律法规',
                '保存证据方法',
                '明确维权途径'
            ]
        },
        {
            'id': 7,
            'category': 'legal',
            'title': '征信异议申诉法律知识',
            'content': """
尊敬的征信中心/银行客服：

我对贵机构报送的征信记录有异议，具体情况如下：

1. 逾期记录与实际情况不符：[说明具体情况]
2. 金额数据存在错误：[指出具体错误]
3. 非本人主观意愿造成的逾期：[说明原因，如银行系统问题等]

**法律依据：**
1. 《征信业管理条例》第25条：信息主体有权提出异议并要求更正
2. 《民法典》第1029条：信用评价错误的更正权
3. 错误征信信息应在20日内核查并更正

**申诉材料：**
1. 征信异议申请书
2. 身份证明文件
3. 相关证明材料（银行对账单、还款凭证等）
4. 情况说明文件

**申诉流程：**
1. 向征信中心或数据报送机构提出书面异议
2. 机构应在20日内核查
3. 确认错误应更正并书面回复
4. 可向人民银行征信管理部门投诉
            """,
            'difficulty': '中级',
            'usage_count': 430,
            'success_rate': 70,
            'tags': ['征信异议', '法律申诉', '征信修复'],
            'key_points': [
                '明确指出征信问题',
                '引用具体法规条款',
                '提供完整申诉材料',
                '了解申诉流程'
            ]
        },
        {
            'id': 8,
            'category': 'legal',
            'title': '债务重组与破产法律知识',
            'content': """
**个人债务重组法律知识：**
1. 可以与多个债权人协商，制定统一的还款计划
2. 可通过债务重组服务机构协助
3. 重组协议需所有债权人同意

**个人破产制度（试点地区）：**
1. 深圳、浙江等试点地区已实施个人破产条例
2. 符合条件的债务人可申请破产保护
3. 经过3-5年考察期可免除剩余债务

**法律注意事项：**
1. 债务重组需专业法律咨询
2. 破产申请有严格条件限制
3. 需如实申报所有财产和债务
4. 恶意逃债将承担法律责任

**适用情况：**
1. 多笔债务无法偿还
2. 有还款意愿但无还款能力
3. 经过专业评估确需法律保护
            """,
            'difficulty': '高级',
            'usage_count': 210,
            'success_rate': 60,
            'tags': ['债务重组', '个人破产', '法律程序'],
            'key_points': [
                '了解债务重组流程',
                '知晓个人破产制度',
                '评估适用条件',
                '寻求专业法律帮助'
            ]
        },
        
        # ===== 心理技巧经验（新增） =====
        {
            'id': 9,
            'category': 'psychological',
            'title': '催收电话心理应对经验',
            'content': """
**接听催收电话心理准备：**
1. 深呼吸，保持平和心态
2. 明确沟通目标：协商还款方案
3. 不被对方情绪影响，专注于解决问题
4. 记录关键信息：对方工号、承诺内容、时间等
5. 适时结束不愉快的对话，换个时间再联系

**有效沟通话术经验：**
- "我理解您的工作职责，但也请您理解我的实际困难。"
- "我们能不能一起找个双方都能接受的解决方案？"
- "刚才您承诺的[内容]，我会记录下来，希望您也能履行承诺。"
- "我会在[时间]前给您回复，请给我一些时间处理。"

**情绪管理技巧：**
1. 设定每天接听电话的时间段
2. 不接陌生号码，通过短信或微信文字沟通
3. 准备一个"压力释放"活动（如散步、听音乐）
4. 与支持你的人分享感受
            """,
            'difficulty': '初级',
            'usage_count': 1560,
            'success_rate': 90,
            'tags': ['情绪管理', '沟通技巧', '心理调节'],
            'key_points': [
                '保持冷静和专业',
                '明确沟通目标',
                '记录关键信息',
                '使用建设性语言'
            ]
        },
        {
            'id': 10,
            'category': 'psychological',
            'title': '债务压力心理调适经验',
            'content': """
**认知重构经验分享：**
1. 债务是暂时困难，不是人生终点
2. 很多成功人士也曾经历过债务困境
3. 还清债务后，你会更懂财务管理
4. 这段经历会成为你人生的宝贵财富

**压力分解方法：**
1. 将大目标分解为可执行的小步骤
2. 每天完成一个小任务，获得成就感
3. 制作还款进度表，可视化进步
4. 庆祝每一个小胜利

**支持系统建立：**
1. 与理解你的家人朋友沟通
2. 加入债务管理经验分享社区
3. 寻找正能量的榜样和故事
4. 避免与消极的人过多交流

**每日心理练习：**
1. 早晨：写下今天要完成的一件事
2. 中午：深呼吸5分钟，放松身心
3. 晚上：记录今天的进步和感恩的事
            """,
            'difficulty': '中级',
            'usage_count': 890,
            'success_rate': 85,
            'tags': ['压力管理', '认知重构', '支持系统'],
            'key_points': [
                '改变对债务的认知',
                '分解压力为小步骤',
                '建立支持系统',
                '坚持每日心理练习'
            ]
        },
        {
            'id': 11,
            'category': 'psychological',
            'title': '长期心理韧性培养经验',
            'content': """
**心理韧性培养经验：**
1. 接受现实，停止自责和内耗
2. 专注于你能控制的事情
3. 从挫折中学习，不重复同样的错误
4. 保持希望，相信未来会更好

**应对焦虑的方法：**
1. 4-7-8呼吸法：吸气4秒，屏气7秒，呼气8秒
2. 5-4-3-2-1感官法：找到5个看到的东西、4个触摸到的东西等
3. 渐进式肌肉放松法
4. 正念冥想练习

**建立积极心态：**
1. 每天记录3件好事
2. 练习感恩，写下感恩清单
3. 设定现实可行的目标
4. 奖励自己的每一个进步

**预防抑郁的日常习惯：**
1. 规律作息，保证充足睡眠
2. 每天适量运动（如散步30分钟）
3. 健康饮食，避免酒精和过量咖啡因
4. 保持社交活动，不自我孤立
            """,
            'difficulty': '高级',
            'usage_count': 520,
            'success_rate': 88,
            'tags': ['心理韧性', '焦虑应对', '积极心态'],
            'key_points': [
                '培养心理韧性',
                '学习焦虑应对技巧',
                '建立积极心态',
                '养成健康生活习惯'
            ]
        }
    ],
    'learning_path': [
        {
            'step': 1,
            'title': '基础认知',
            'description': '了解协商基本原则和准备工作经验',
            'duration': '20分钟',
            'phrases': [1, 9]  # 信用卡逾期协商 + 催收电话心理应对
        },
        {
            'step': 2,
            'title': '银行协商',
            'description': '交流银行信用卡和贷款协商技巧经验',
            'duration': '30分钟',
            'phrases': [2, 6]  # 个性化分期 + 违规催收应对
        },
        {
            'step': 3,
            'title': '网贷协商',
            'description': '学习网贷平台协商策略经验',
            'duration': '25分钟',
            'phrases': [3, 4]  # 延期还款 + 高利率协商
        },
        {
            'step': 4,
            'title': '法律知识',
            'description': '运用法律知识保护合法权益',
            'duration': '35分钟',
            'phrases': [7, 8]  # 征信异议 + 债务重组知识
        },
        {
            'step': 5,
            'title': '心理调适',
            'description': '掌握债务压力下的心理调适方法',
            'duration': '30分钟',
            'phrases': [10, 11]  # 债务压力调适 + 心理韧性培养
        },
        {
            'step': 6,
            'title': '实战进阶',
            'description': '高级协商技巧和综合应对策略',
            'duration': '40分钟',
            'phrases': [5]  # 一次性结清协商
        }
    ]
}

@app.route('/negotiation-guide')
@payment_required
def negotiation_guide():
    """协商经验分享大全主页面"""
    return render_template('negotiation_guide.html')

@app.route('/api/negotiation-phrases')
@payment_required
def get_negotiation_phrases():
    """获取协商经验分享数据"""
    try:
        # 计算热门话术（按使用次数排序）
        popular_phrases = sorted(
            [p for p in NEGOTIATION_PHRASES['phrases']],
            key=lambda x: x['usage_count'],
            reverse=True
        )[:6]
        
        return jsonify({
            'success': True,
            'categories': NEGOTIATION_PHRASES['categories'],
            'phrases': NEGOTIATION_PHRASES['phrases'],
            'learning_path': NEGOTIATION_PHRASES['learning_path'],
            'popular_phrases': popular_phrases,
            'stats': {
                'total_phrases': len(NEGOTIATION_PHRASES['phrases']),
                'total_categories': len(NEGOTIATION_PHRASES['categories']),
                'avg_success_rate': sum(p['success_rate'] for p in NEGOTIATION_PHRASES['phrases']) // len(NEGOTIATION_PHRASES['phrases'])
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取经验数据失败: {str(e)}'})

@app.route('/api/negotiation-phrase/<int:phrase_id>')
@payment_required
def get_negotiation_phrase(phrase_id):
    """获取特定经验分享的详细信息"""
    try:
        phrase = next((p for p in NEGOTIATION_PHRASES['phrases'] if p['id'] == phrase_id), None)
        if phrase:
            # 增加使用计数（在实际应用中应该持久化到数据库）
            phrase['usage_count'] += 1
            
            # 获取相关经验分享推荐
            category_phrases = [p for p in NEGOTIATION_PHRASES['phrases'] 
                              if p['category'] == phrase['category'] and p['id'] != phrase_id][:3]
            
            return jsonify({
                'success': True,
                'phrase': phrase,
                'related_phrases': category_phrases
            })
        else:
            return jsonify({'success': False, 'message': '经验分享不存在'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取经验详情失败: {str(e)}'})

@app.route('/api/negotiation-category/<category_id>')
@payment_required
def get_negotiation_category(category_id):
    """获取特定分类的经验分享"""
    try:
        category = next((c for c in NEGOTIATION_PHRASES['categories'] if c['id'] == category_id), None)
        if category:
            category_phrases = [p for p in NEGOTIATION_PHRASES['phrases'] if p['category'] == category_id]
            return jsonify({
                'success': True,
                'category': category,
                'phrases': category_phrases
            })
        else:
            return jsonify({'success': False, 'message': '分类不存在'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取分类经验失败: {str(e)}'})

# ============ 启动应用 ============
if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)